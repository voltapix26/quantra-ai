# Quantra AI — investor-grade financial model.
# Budget-driven (marketing spend -> CAC -> signups), Bear/Base/Bull scenario switch,
# real unit economics (CAC, LTV, LTV:CAC, payback), charts. All live Excel formulas.
# PROJECTIONS ONLY — Quantra has no revenue and no users to date.
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

OUT_DIR = os.path.join(os.path.expanduser('~'), 'Desktop', 'Quantra AI')
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, 'Quantra_AI_Financial_Model.xlsx')

DARK = 'FF0A0F1C'; LIGHT = 'FFEEF2F7'; INPUT = 'FFFFF9E6'; ACT = 'FFE8FBF3'
H = Font(bold=True, color='FFFFFFFF', size=11); B = Font(bold=True)
SM = Font(size=9, italic=True, color='FF667085'); RED = Font(bold=True, size=9, color='FFB00020')
fillH = PatternFill('solid', fgColor=DARK); fillS = PatternFill('solid', fgColor=LIGHT)
fillIn = PatternFill('solid', fgColor=INPUT); fillAct = PatternFill('solid', fgColor=ACT)
thin = Border(bottom=Side(style='thin', color='FFDDDDDD'))

wb = Workbook()

# ================= ASSUMPTIONS (scenario-driven) =================
a = wb.active; a.title = 'Assumptions'
a['A1'] = 'QUANTRA AI — ASSUMPTIONS & SCENARIOS'; a['A1'].font = Font(bold=True, size=14)
a['A2'] = 'SCENARIO (1=Bear, 2=Base, 3=Bull) →'; a['A2'].font = B
a['B2'] = 2; a['B2'].fill = fillIn; a['B2'].font = Font(bold=True, size=13); a['B2'].alignment = Alignment(horizontal='center')
a['C2'] = '=CHOOSE($B$2,"BEAR","BASE","BULL")'; a['C2'].font = Font(bold=True, size=13, color='FF0E9F6E')
a['A3'] = 'Change B2 to flip the whole model. Yellow = inputs. Green = the active value feeding the model.'; a['A3'].font = SM
a['A4'] = 'PROJECTIONS ONLY — Quantra has no revenue and no active users to date. No historical data in this model.'; a['A4'].font = RED

hdr = ['Driver', 'Bear', 'Base', 'Bull', 'ACTIVE', 'Note']
for i, h in enumerate(hdr):
    c = a.cell(6, i + 1, h); c.font = H; c.fill = fillH; c.alignment = Alignment(horizontal='center')

# (label, bear, base, bull, fmt, note)  — fmt: p=percent, m=money, n=number
D = [
    ('ACQUISITION', None, None, None, None, None),
    ('Marketing spend — Y1 (US$/mo)', 8000, 12000, 15000, 'm', 'A decision, not a guess. Signups follow from CAC.'),
    ('Marketing spend — Y2 (US$/mo)', 20000, 30000, 40000, 'm', 'Scales only after CAC is proven'),
    ('Marketing spend — Y3 (US$/mo)', 30000, 50000, 65000, 'm', ''),
    ('Paid CAC per signup — Y1 (US$)', 9.0, 6.0, 4.5, 'm', 'Blended GCC (higher) + India (lower) performance marketing'),
    ('Paid CAC per signup — Y2 (US$)', 11.0, 7.5, 5.5, 'm', 'CAC inflates as the cheap audience saturates'),
    ('Paid CAC per signup — Y3 (US$)', 13.0, 9.0, 6.5, 'm', ''),
    ('Organic signups (% of paid)', 0.15, 0.35, 0.60, 'p', 'Share-loop cards, SEO, word of mouth — CAC-free. Loop is live in product.'),
    ('MONETISATION', None, None, None, None, None),
    ('Free → paid conversion', 0.025, 0.04, 0.06, 'p', 'Prosumer fintech range 2-6%'),
    ('Pro price (US$/mo)', 9.99, 9.99, 12.99, 'm', 'Plan already built in product'),
    ('Ultimate price (US$/mo)', 24.99, 24.99, 29.99, 'm', 'Plan already built in product'),
    ('Mix — share on Pro', 0.88, 0.80, 0.72, 'p', 'Remainder on Ultimate'),
    ('Monthly churn — paying', 0.07, 0.05, 0.035, 'p', 'Drives LTV directly'),
    ('Monthly churn — free', 0.10, 0.08, 0.06, 'p', ''),
    ('Payment processing fee', 0.03, 0.03, 0.03, 'p', 'Stripe + FX'),
    ('COSTS (US$/mo)', None, None, None, None, None),
    ('Salaries — Y1', 30000, 26000, 26000, 'm', 'Founder + first hires (eng, data, growth)'),
    ('Salaries — Y2', 58000, 52000, 52000, 'm', 'Team ~7'),
    ('Salaries — Y3', 80000, 72000, 72000, 'm', 'Team ~10'),
    ('Market data & licences — Y1', 2500, 2000, 2000, 'm', 'Paid API tiers + first exchange licence'),
    ('Market data & licences — Y2', 7000, 6000, 6000, 'm', 'Gulf + NSE/BSE licensed feeds'),
    ('Market data & licences — Y3', 12000, 10000, 10000, 'm', ''),
    ('Infrastructure — Y1', 1000, 800, 800, 'm', 'Paid instance, DB, CDN, AI API'),
    ('Infrastructure — Y2', 3500, 3000, 3000, 'm', ''),
    ('Infrastructure — Y3', 8000, 7000, 7000, 'm', ''),
    ('G&A, legal & compliance — Y1', 7000, 6000, 6000, 'm', 'ADGM/UAE path, audit, insurance'),
    ('G&A, legal & compliance — Y2', 9000, 8000, 8000, 'm', ''),
    ('G&A, legal & compliance — Y3', 11000, 10000, 10000, 'm', ''),
    ('FUNDING', None, None, None, None, None),
    ('Seed raise (US$)', 2500000, 2500000, 2500000, 'm', 'This round'),
]
r = 7; K = {}
for label, bear, base, bull, fmt, note in D:
    if bear is None:
        for c in range(1, 7):
            a.cell(r, c).fill = fillH
        a.cell(r, 1, label).font = H
    else:
        a.cell(r, 1, label).font = B
        for i, v in enumerate([bear, base, bull]):
            c = a.cell(r, 2 + i, v); c.fill = fillIn
            c.number_format = '0.0%' if fmt == 'p' else ('#,##0.00' if v < 100 else '#,##0')
        act = a.cell(r, 5, f'=CHOOSE($B$2,B{r},C{r},D{r})')
        act.fill = fillAct; act.font = B
        act.number_format = '0.0%' if fmt == 'p' else ('#,##0.00' if base < 100 else '#,##0')
        a.cell(r, 6, note).font = SM
        K[label] = f'Assumptions!$E${r}'
    r += 1
a.column_dimensions['A'].width = 34
for col, w in zip('BCDE', [11, 11, 11, 13]): a.column_dimensions[col].width = w
a.column_dimensions['F'].width = 56

# ================= MONTHLY MODEL =================
m = wb.create_sheet('Monthly Model')
m['A1'] = 'MONTHLY MODEL — 36 MONTHS (every cell a formula; month 1 = first month post-funding)'
m['A1'].font = Font(bold=True, size=13)
LAB = {
    3: 'Month',
    4: 'Marketing spend',
    5: 'Paid signups', 6: 'Organic signups', 7: 'Total new signups',
    8: 'Active users (after churn)', 9: 'New paying customers', 10: 'Paying customers (after churn)',
    11: 'Blended ARPU', 12: 'MRR', 13: 'Revenue',
    15: 'Salaries', 16: 'Market data & licences', 17: 'Infrastructure', 18: 'Marketing (=row 4)',
    19: 'G&A / legal', 20: 'Payment fees', 21: 'Total expense', 22: 'EBITDA', 23: 'Cash balance',
    25: 'CAC per paying customer', 26: 'LTV per paying customer', 27: 'LTV : CAC', 28: 'CAC payback (months)',
}
for row, lab in LAB.items():
    c = m.cell(row, 1, lab); c.font = B
    if row in (13, 21, 22, 23, 27): c.fill = fillS
m.column_dimensions['A'].width = 26

for i in range(1, 37):
    L = get_column_letter(1 + i); prev = get_column_letter(i) if i > 1 else None
    m[f'{L}3'] = i; m[f'{L}3'].font = B; m[f'{L}3'].alignment = Alignment(horizontal='center')
    def yr(k1, k2, k3): return f"=IF({L}$3<=12,{K[k1]},IF({L}$3<=24,{K[k2]},{K[k3]}))"
    m[f'{L}4'] = yr('Marketing spend — Y1 (US$/mo)', 'Marketing spend — Y2 (US$/mo)', 'Marketing spend — Y3 (US$/mo)')
    cac = f"IF({L}$3<=12,{K['Paid CAC per signup — Y1 (US$)']},IF({L}$3<=24,{K['Paid CAC per signup — Y2 (US$)']},{K['Paid CAC per signup — Y3 (US$)']}))"
    m[f'{L}5'] = f"={L}4/{cac}"
    m[f'{L}6'] = f"={L}5*{K['Organic signups (% of paid)']}"
    m[f'{L}7'] = f"={L}5+{L}6"
    m[f'{L}8'] = (f"={L}7" if i == 1 else f"={prev}8*(1-{K['Monthly churn — free']})+{L}7")
    m[f'{L}9'] = f"={L}7*{K['Free → paid conversion']}"
    m[f'{L}10'] = (f"={L}9" if i == 1 else f"={prev}10*(1-{K['Monthly churn — paying']})+{L}9")
    m[f'{L}11'] = (f"={K['Pro price (US$/mo)']}*{K['Mix — share on Pro']}"
                   f"+{K['Ultimate price (US$/mo)']}*(1-{K['Mix — share on Pro']})")
    m[f'{L}12'] = f"={L}10*{L}11"
    m[f'{L}13'] = f"={L}12"
    m[f'{L}15'] = yr('Salaries — Y1', 'Salaries — Y2', 'Salaries — Y3')
    m[f'{L}16'] = yr('Market data & licences — Y1', 'Market data & licences — Y2', 'Market data & licences — Y3')
    m[f'{L}17'] = yr('Infrastructure — Y1', 'Infrastructure — Y2', 'Infrastructure — Y3')
    m[f'{L}18'] = f"={L}4"
    m[f'{L}19'] = yr('G&A, legal & compliance — Y1', 'G&A, legal & compliance — Y2', 'G&A, legal & compliance — Y3')
    m[f'{L}20'] = f"={L}13*{K['Payment processing fee']}"
    m[f'{L}21'] = f"=SUM({L}15:{L}20)"
    m[f'{L}22'] = f"={L}13-{L}21"
    m[f'{L}23'] = (f"={K['Seed raise (US$)']}+{L}22" if i == 1 else f"={prev}23+{L}22")
    # unit economics
    m[f'{L}25'] = f"=IFERROR({L}4/{L}9,0)"
    m[f'{L}26'] = f"=IFERROR({L}11*(1-{K['Payment processing fee']})/{K['Monthly churn — paying']},0)"
    m[f'{L}27'] = f"=IFERROR({L}26/{L}25,0)"
    m[f'{L}28'] = f"=IFERROR({L}25/({L}11*(1-{K['Payment processing fee']})),0)"
    for row in (4, 5, 6, 7, 8, 9, 10, 12, 13, 15, 16, 17, 18, 19, 20, 21, 22, 23):
        m[f'{L}{row}'].number_format = '#,##0'
    for row in (11, 25, 26, 27, 28): m[f'{L}{row}'].number_format = '0.00'
    m.column_dimensions[L].width = 11

# ================= ANNUAL SUMMARY =================
s = wb.create_sheet('Annual Summary', 0)
s['A1'] = 'QUANTRA AI — 3-YEAR PROJECTION'; s['A1'].font = Font(bold=True, size=16)
s['A2'] = '=\"Scenario: \"&CHOOSE(Assumptions!$B$2,\"BEAR\",\"BASE\",\"BULL\")&\"  —  change Assumptions!B2 to switch\"'
s['A2'].font = Font(bold=True, size=11, color='FF0E9F6E')
s['A3'] = 'Projections only — Quantra has no revenue and no active users to date. Every figure is a formula off the Assumptions sheet.'
s['A3'].font = RED
for i, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3']):
    c = s.cell(5, i + 1, h); c.font = H; c.fill = fillH; c.alignment = Alignment(horizontal='center')

def rng(row, y):
    return f"'Monthly Model'!{get_column_letter(2 + (y - 1) * 12)}{row}:{get_column_letter(1 + y * 12)}{row}"
def end(y): return get_column_letter(1 + y * 12)

METRICS = [
    ('HEADLINE', None, None),
    ('Total Customers (paying, end of year)', lambda y: f"='Monthly Model'!{end(y)}10", '#,##0'),
    ('Total Revenue (US$)', lambda y: f"=SUM({rng(13, y)})", '#,##0'),
    ('Total Expense (US$)', lambda y: f"=SUM({rng(21, y)})", '#,##0'),
    ('EBITDA (US$)', lambda y: f"=SUM({rng(22, y)})", '#,##0'),
    ('GROWTH', None, None),
    ('Active users (end of year)', lambda y: f"='Monthly Model'!{end(y)}8", '#,##0'),
    ('New signups in year', lambda y: f"=SUM({rng(7, y)})", '#,##0'),
    ('Exit MRR (US$)', lambda y: f"='Monthly Model'!{end(y)}12", '#,##0'),
    ('ARR run-rate (US$)', lambda y: f"='Monthly Model'!{end(y)}12*12", '#,##0'),
    ('Revenue growth', lambda y: '—' if y == 1 else f"=IFERROR(SUM({rng(13, y)})/SUM({rng(13, y-1)})-1,0)", '0%'),
    ('UNIT ECONOMICS', None, None),
    ('CAC per paying customer (US$)', lambda y: f"=IFERROR(SUM({rng(4, y)})/SUM({rng(9, y)}),0)", '0.00'),
    ('LTV per paying customer (US$)', lambda y: f"='Monthly Model'!{end(y)}26", '0.00'),
    ('LTV : CAC', lambda y: f"=IFERROR('Monthly Model'!{end(y)}26/(SUM({rng(4, y)})/SUM({rng(9, y)})),0)", '0.00'),
    ('CAC payback (months)', lambda y: f"='Monthly Model'!{end(y)}28", '0.0'),
    ('CASH', None, None),
    ('EBITDA margin', lambda y: f"=IFERROR(SUM({rng(22, y)})/SUM({rng(13, y)}),0)", '0%'),
    ('Cash balance (end of year, US$)', lambda y: f"='Monthly Model'!{end(y)}23", '#,##0'),
]
r = 6
for label, f, fmt in METRICS:
    if f is None:
        for c in range(1, 5): s.cell(r, c).fill = fillH
        s.cell(r, 1, label).font = H
    else:
        c = s.cell(r, 1, label); c.font = B; c.border = thin
        for y in (1, 2, 3):
            cc = s.cell(r, 1 + y, f(y)); cc.number_format = fmt; cc.border = thin
    r += 1
s.column_dimensions['A'].width = 34
for col in 'BCD': s.column_dimensions[col].width = 15

r += 1
s.cell(r, 1, 'HOW TO READ THIS MODEL').font = Font(bold=True, size=12); r += 1
for n in [
    'Budget-driven, not wish-driven: marketing spend is an input; signups fall out of it at the stated CAC. Growth is never assumed — it is bought at a price we state, plus an organic share from the in-product share loop (live today).',
    'UNIT ECONOMICS, STATED HONESTLY: the Base case runs at roughly 2.3x LTV:CAC in year 1 and drifts to ~1.5x by year 3 as CAC inflates. That is BELOW the 3:1 bar investors look for. We have not tuned the assumptions until 3:1 appeared.',
    'Why it is below, and what we do about it: CAC and conversion are the two numbers a pre-launch company cannot know — we have no users yet. The first milestone of this round is to MEASURE both on a small cohort (~US$25k of spend) before scaling. Marketing only scales once the ratio clears 3:1. If it cannot be made to clear, the honest outcome is a smaller, slower company — not a bigger burn.',
    'The levers that move it, in order: conversion (4% → 5% takes Base to ~2.8x), organic share from the share loop (35% → 60%), CAC (Gulf is dear, India is cheap — mix matters), and churn (drives LTV directly). Bull shows all four landing well: 10.5x. Bear shows them missing: 0.4x.',
    'Runway: the US$2.5M seed funds the plan into year 3. Closing year-3 cash is negative in Base — this plan assumes a Series A in year 3 against a ~US$1M ARR run-rate. Without it, marketing and hiring throttle back to hold cash positive; growth slows, the company survives.',
    'EBITDA is negative across all three years. This is a growth-stage SaaS plan; breakeven is not claimed inside the seed window.',
    'Not modelled (all upside, none relied on): developer API revenue, team/enterprise seats, licensed-data resale, any exit.',
]:
    s.cell(r, 1, '•  ' + n).font = Font(size=9.5); r += 1

r += 1
s.cell(r, 1, 'USE OF FUNDS — US$2.5M').font = Font(bold=True, size=12); r += 1
for label, pct, amt, what in [
    ('Team', '35%', 875000, '4 engineers + data specialist + growth lead; 24-month runway'),
    ('Data & infrastructure', '30%', 750000, 'Licensed Gulf + NSE/BSE feeds, premium APIs, production hosting'),
    ('Growth & go-to-market', '20%', 500000, 'GCC + India marketing, app-store launches, referral loops'),
    ('Regulatory & legal', '10%', 250000, 'ADGM/UAE footing for a paid research product'),
    ('Reserve', '5%', 125000, 'Contingency, audit, insurance'),
]:
    s.cell(r, 1, label).font = B
    s.cell(r, 2, pct).alignment = Alignment(horizontal='center')
    c = s.cell(r, 3, amt); c.number_format = '#,##0'
    s.cell(r, 4, what).font = SM
    r += 1
r += 1
s.cell(r, 1, 'Forward-looking projections on the stated assumptions — not forecasts or guarantees. Actual results will differ.').font = RED

# ================= CHARTS =================
ch = wb.create_sheet('Charts')
ch['A1'] = 'CHARTS — driven by the Monthly Model (Base/Bear/Bull switch applies)'; ch['A1'].font = Font(bold=True, size=13)
ch['A3'] = 'Month'; ch['B3'] = 'Revenue'; ch['C3'] = 'Total expense'; ch['D3'] = 'Paying customers'; ch['E3'] = 'Cash balance'
for i in range(1, 37):
    L = get_column_letter(1 + i)
    ch.cell(3 + i, 1, i)
    ch.cell(3 + i, 2, f"='Monthly Model'!{L}13")
    ch.cell(3 + i, 3, f"='Monthly Model'!{L}21")
    ch.cell(3 + i, 4, f"='Monthly Model'!{L}10")
    ch.cell(3 + i, 5, f"='Monthly Model'!{L}23")

c1 = LineChart(); c1.title = 'Revenue vs Expense (US$/month)'; c1.height = 8; c1.width = 17
c1.add_data(Reference(ch, min_col=2, min_row=3, max_col=3, max_row=39), titles_from_data=True)
c1.set_categories(Reference(ch, min_col=1, min_row=4, max_row=39))
ch.add_chart(c1, 'G3')
c2 = BarChart(); c2.title = 'Paying customers'; c2.height = 8; c2.width = 17
c2.add_data(Reference(ch, min_col=4, min_row=3, max_row=39), titles_from_data=True)
c2.set_categories(Reference(ch, min_col=1, min_row=4, max_row=39))
ch.add_chart(c2, 'G20')
c3 = LineChart(); c3.title = 'Cash balance (runway)'; c3.height = 8; c3.width = 17
c3.add_data(Reference(ch, min_col=5, min_row=3, max_row=39), titles_from_data=True)
c3.set_categories(Reference(ch, min_col=1, min_row=4, max_row=39))
ch.add_chart(c3, 'G37')

wb.save(OUT)
print('SAVED:', OUT)
